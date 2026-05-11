"""
=========================================================
FBA Inventory Intelligence Dashboard
=========================================================
Backend: Flask + Pandas
Purpose: Detect excess inventory being sent to Amazon FBA
         by analyzing shipments, sales, and current stock.
=========================================================
"""

import csv
import os
import math
import uuid
from datetime import datetime, timedelta
from io import BytesIO, TextIOWrapper

import numpy as np
import pandas as pd
from flask import (
    Flask, render_template, request, jsonify, send_file, abort
)
from openpyxl import load_workbook
from werkzeug.utils import secure_filename


# ----------------------------------------------------------
# Flask app configuration
# ----------------------------------------------------------
app = Flask(__name__)

# Folder paths
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, "static", "uploads")
OUTPUT_FOLDER = os.path.join(BASE_DIR, "outputs")

# Make sure folders exist
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50 MB max upload

# Allowed file types
ALLOWED_EXTENSIONS = {"xlsx", "xls", "csv"}

# Required columns for each file type
REQUIRED_COLUMNS = {
    "shipment": ["Loading Date", "Appointment ID", "Shipment ID", "SKU", "Qty"],
    "sales":    ["ASIN", "SKU", "FNSKU", "Qty"],
    "stock":    ["SKU", "ASIN", "FNSKU", "Qty"],
}

# In-memory store of the last computed dashboard (so download endpoints
# can re-use the result without recomputing). Keyed by session_id.
LAST_RESULT = {}


@app.after_request
def add_cors_headers(response):
    response.headers["Access-Control-Allow-Origin"] = "*"
    response.headers["Access-Control-Allow-Headers"] = "Content-Type, Authorization"
    response.headers["Access-Control-Allow-Methods"] = "GET, POST, OPTIONS"
    return response


# ----------------------------------------------------------
# Helper functions
# ----------------------------------------------------------
def allowed_file(filename: str) -> bool:
    """Check if uploaded file has a supported extension."""
    return (
        "." in filename
        and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS
    )


def read_csv_limited(stream, required_columns):
    stream.seek(0)
    wrapper = TextIOWrapper(stream, encoding="utf-8", errors="replace")
    reader = csv.DictReader(wrapper)
    rows = []
    for row in reader:
        rows.append({col: row.get(col, None) for col in required_columns})
    wrapper.detach()
    df = pd.DataFrame(rows, columns=required_columns)
    if "Qty" in df.columns:
        df["Qty"] = pd.to_numeric(df["Qty"], errors="coerce").fillna(0).astype(int)
    return df


def read_xlsx_limited(stream, required_columns):
    stream.seek(0)
    try:
        df = pd.read_excel(
            stream,
            engine="openpyxl",
            usecols=lambda col: str(col).strip() in required_columns,
            dtype={"Qty": "Int64"},
        )
    except Exception:
        stream.seek(0)
        wb = load_workbook(stream, read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        try:
            header = [str(cell).strip() if cell is not None else "" for cell in next(rows)]
        except StopIteration:
            return pd.DataFrame(columns=required_columns)

        selected = [i for i, name in enumerate(header) if name in required_columns]
        if not selected:
            return pd.DataFrame(columns=required_columns)

        selected_names = [header[i] for i in selected]
        data = []
        for row in rows:
            data.append({
                selected_names[j]: row[idx] if idx < len(row) else None
                for j, idx in enumerate(selected)
            })
        df = pd.DataFrame(data, columns=selected_names)

    df.columns = [str(c).strip() for c in df.columns]
    if "Qty" in df.columns:
        df["Qty"] = pd.to_numeric(df["Qty"], errors="coerce").fillna(0).astype(int)
    return df


def read_any(file_storage, required_columns) -> pd.DataFrame:
    """
    Read an uploaded file (xlsx / xls / csv) into a Pandas DataFrame.
    Works with werkzeug FileStorage objects directly (no disk needed).
    """
    filename = file_storage.filename.lower()
    file_storage.stream.seek(0)

    if filename.endswith(".csv"):
        return read_csv_limited(file_storage.stream, required_columns)
    if filename.endswith(".xls"):
        return pd.read_excel(
            file_storage.stream,
            engine="xlrd",
            usecols=lambda col: str(col).strip() in required_columns,
        )
    return read_xlsx_limited(file_storage.stream, required_columns)


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Strip whitespace from column names for safer matching."""
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    return df


def validate_columns(df: pd.DataFrame, required: list, label: str):
    """Raise ValueError if required columns are missing."""
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(
            f"{label} file is missing required columns: {', '.join(missing)}"
        )


def safe_str(x) -> str:
    """Convert any value to a clean string (NaN/None -> '')."""
    if pd.isna(x):
        return ""
    return str(x).strip()


def get_status(daily_sales: float, days_cover: float) -> str:
    """
    Apply business rules to classify each SKU.
    Order matters — the first matching condition wins.
    """
    if daily_sales == 0:
        return "Dead Stock"
    if days_cover < 10:
        return "Critical"
    if days_cover < 20:
        return "Low Stock"
    if days_cover <= 45:
        return "Healthy"
    if days_cover <= 60:
        return "Overstock"
    return "Excess Inventory"


def normalize_day(value=None) -> pd.Timestamp:
    """Return a normalized date for inventory projections."""
    if value is None:
        return pd.Timestamp.now().normalize()
    return pd.Timestamp(value).normalize()


def build_events_by_sku(shipment_df: pd.DataFrame, lead_time: int) -> dict:
    """Build ordered shipment arrival events for each SKU."""
    schedule = shipment_df.copy()
    schedule["Qty"] = pd.to_numeric(schedule["Qty"], errors="coerce").fillna(0)
    schedule["Loading Date Parsed"] = schedule["Loading Date"].apply(parse_date)
    schedule["Arrival Date"] = schedule["Loading Date Parsed"].apply(
        lambda dt: dt + timedelta(days=lead_time) if dt is not None else None
    )

    events_by_sku = {}
    for _, row in schedule.iterrows():
        sku = safe_str(row["SKU"])
        if not sku:
            continue

        qty = float(row["Qty"]) if pd.notna(row["Qty"]) else 0.0
        events_by_sku.setdefault(sku, []).append((row["Arrival Date"], qty))

    def event_sort_key(event):
        arrival_date = event[0]
        if arrival_date is None or pd.isna(arrival_date):
            return pd.Timestamp.max
        return pd.Timestamp(arrival_date).normalize()

    for sku in events_by_sku:
        events_by_sku[sku].sort(key=event_sort_key)

    return events_by_sku


def project_stock_through_events(
    current_stock: float,
    daily_sales: float,
    events: list,
    today: pd.Timestamp,
    fallback_days: int = 0,
) -> tuple[float, list]:
    """
    Project stock by first consuming sales until each shipment arrival, then
    adding the arriving quantity.
    """
    stock = max(0.0, float(current_stock))
    sales_rate = max(0.0, float(daily_sales))
    current_day = normalize_day(today)

    if not events:
        days = max(0, int(fallback_days))
        return max(0.0, stock - (sales_rate * days)), []

    snapshots = []
    for arrival_date, qty in events:
        arrival_day = (
            pd.Timestamp(arrival_date).normalize()
            if arrival_date is not None and not pd.isna(arrival_date)
            else current_day
        )
        days_until_arrival = max(0, (arrival_day - current_day).days)
        consumed = sales_rate * days_until_arrival
        stock_before_arrival = max(0.0, stock - consumed)
        shipment_qty = float(qty) if pd.notna(qty) else 0.0
        stock_after_arrival = stock_before_arrival + shipment_qty

        snapshots.append({
            "arrival_date": arrival_day,
            "days_until_arrival": days_until_arrival,
            "consumed": consumed,
            "stock_before_arrival": stock_before_arrival,
            "shipment_qty": shipment_qty,
            "stock_after_arrival": stock_after_arrival,
        })

        stock = stock_after_arrival
        current_day = arrival_day

    return stock, snapshots


# ----------------------------------------------------------
# Core calculation engine
# ----------------------------------------------------------
def build_dashboard(
    shipment_df: pd.DataFrame,
    sales_df: pd.DataFrame,
    stock_df: pd.DataFrame,
    transit_days: int,
    grn_days: int,
    target_days: int,
    today=None,
) -> pd.DataFrame:
    """
    Merge the three input files and compute every metric required
    by the dashboard. Returns a single DataFrame keyed by SKU.
    """
    lead_time = transit_days + grn_days
    today = normalize_day(today)

    # --- 1. Aggregate shipment quantities per SKU -----------------------
    ship = (
        shipment_df.groupby("SKU", as_index=False)["Qty"]
        .sum()
        .rename(columns={"Qty": "Incoming Shipment Qty"})
    )

    # --- 1a. Build arrival schedule for future shipments ---------------
    events_by_sku = build_events_by_sku(shipment_df, lead_time)

    def project_stock_after_arrivals(row):
        events = events_by_sku.get(safe_str(row["SKU"]), [])
        projected_stock, _ = project_stock_through_events(
            row["Current FBA Stock"],
            row["Daily Sales"],
            events,
            today,
            fallback_days=lead_time,
        )
        return projected_stock

    # --- 2. Aggregate 45-day sales per SKU ------------------------------
    sales = (
        sales_df.groupby(["SKU", "ASIN", "FNSKU"], as_index=False, dropna=False)["Qty"]
        .sum()
        .rename(columns={"Qty": "45 Days Sales"})
    )

    # --- 3. Aggregate current FBA stock per SKU -------------------------
    stock = (
        stock_df.groupby(["SKU", "ASIN", "FNSKU"], as_index=False, dropna=False)["Qty"]
        .sum()
        .rename(columns={"Qty": "Current FBA Stock"})
    )

    # --- 4. Build master SKU list (union of all three sources) ----------
    #     Outer merge keeps every SKU that appears anywhere.
    master = (
        stock.merge(sales, on=["SKU", "ASIN", "FNSKU"], how="outer")
             .merge(ship, on="SKU", how="outer")
    )

    # Fill numeric NaN with 0
    for col in ["45 Days Sales", "Current FBA Stock", "Incoming Shipment Qty"]:
        if col not in master.columns:
            master[col] = 0
        master[col] = pd.to_numeric(master[col], errors="coerce").fillna(0)

    # Fill text NaN with ""
    for col in ["SKU", "ASIN", "FNSKU"]:
        master[col] = master[col].apply(safe_str)

    # Drop rows where SKU is blank (junk)
    master = master[master["SKU"] != ""].copy()

    # --- 5. Per-row calculations ---------------------------------------
    master["Daily Sales"] = (master["45 Days Sales"] / 45).round(2)

    # Projected stock after all scheduled shipments land
    master["Projected Stock"] = (
        master.apply(project_stock_after_arrivals, axis=1)
              .round(0)
              .astype(int)
    )

    def stock_before_first_arrival(row):
        events = events_by_sku.get(safe_str(row["SKU"]), [])
        if not events:
            remaining_stock, _ = project_stock_through_events(
                row["Current FBA Stock"],
                row["Daily Sales"],
                [],
                today,
                fallback_days=lead_time,
            )
            return remaining_stock

        _, snapshots = project_stock_through_events(
            row["Current FBA Stock"],
            row["Daily Sales"],
            events[:1],
            today,
        )
        return snapshots[0]["stock_before_arrival"]

    remaining = master.apply(stock_before_first_arrival, axis=1)
    master["Stock Before Arrival"] = remaining.round(0).astype(int)

    # Required quantity for target days of cover (always round UP)
    master["Required Qty"] = (
        master["Daily Sales"] * target_days
    ).apply(lambda x: int(math.ceil(x)))

    # Days of cover after scheduled shipments have landed.
    master["Days Cover"] = np.where(
        master["Daily Sales"] > 0,
        (master["Projected Stock"] / master["Daily Sales"]).round(1),
        999.0,
    )

    # Excess is only counted after current stock depletion and shipment arrival.
    candidate_excess = (master["Projected Stock"] - master["Required Qty"]).clip(lower=0)
    master["Excess Qty"] = np.where(
        master["Days Cover"] > target_days,
        candidate_excess,
        0,
    ).astype(int)

    # Status classification
    master["Status"] = master.apply(
        lambda r: get_status(r["Daily Sales"], r["Days Cover"]), axis=1
    )

    # Recommended shipment qty = how much we *should* be sending
    master["Recommended Shipment Qty"] = (
        master["Required Qty"] - remaining
    ).clip(lower=0).round(0).astype(int)

    # Extra metadata for export
    master["Lead Time"] = lead_time

    # Clean column ordering
    master = master[[
        "SKU", "ASIN", "FNSKU",
        "45 Days Sales", "Daily Sales",
        "Current FBA Stock", "Incoming Shipment Qty",
        "Lead Time", "Stock Before Arrival", "Projected Stock",
        "Required Qty", "Excess Qty",
        "Days Cover", "Status", "Recommended Shipment Qty",
    ]]

    # Cast integer columns
    int_cols = [
        "45 Days Sales", "Current FBA Stock", "Incoming Shipment Qty",
        "Lead Time", "Stock Before Arrival", "Projected Stock", "Required Qty",
        "Excess Qty", "Recommended Shipment Qty",
    ]
    for c in int_cols:
        master[c] = master[c].astype(int)

    return master.sort_values("Excess Qty", ascending=False).reset_index(drop=True)


def build_kpis(df: pd.DataFrame) -> dict:
    """Compute top-line KPI numbers."""
    return {
        "total_skus":    int(len(df)),
        "healthy":       int((df["Status"] == "Healthy").sum()),
        "critical":      int((df["Status"] == "Critical").sum()),
        "low_stock":     int((df["Status"] == "Low Stock").sum()),
        "overstock":     int((df["Status"] == "Overstock").sum()),
        "excess":        int((df["Status"] == "Excess Inventory").sum()),
        "dead_stock":    int((df["Status"] == "Dead Stock").sum()),
        "total_excess_qty": int(df["Excess Qty"].sum()),
    }


def build_charts(df: pd.DataFrame) -> dict:
    """Build the data payloads Chart.js needs."""
    # Status distribution
    status_counts = df["Status"].value_counts().to_dict()
    status_order = [
        "Healthy", "Low Stock", "Critical",
        "Overstock", "Excess Inventory", "Dead Stock",
    ]
    status_labels = [s for s in status_order if s in status_counts]
    status_values = [int(status_counts[s]) for s in status_labels]

    # Top 10 excess SKUs
    top_excess = df.sort_values("Excess Qty", ascending=False).head(10)
    top_excess = top_excess[top_excess["Excess Qty"] > 0]

    return {
        "status": {
            "labels": status_labels,
            "values": status_values,
        },
        "top_excess": {
            "labels": top_excess["SKU"].tolist(),
            "values": top_excess["Excess Qty"].astype(int).tolist(),
        },
        "stock_health": {
            "labels": ["Projected Stock", "Required Qty", "Excess Qty"],
            "values": [
                int(df["Projected Stock"].sum()),
                int(df["Required Qty"].sum()),
                int(df["Excess Qty"].sum()),
            ],
        },
    }


def parse_date(value):
    if pd.isna(value) or value == "":
        return None
    if isinstance(value, datetime):
        return value
    try:
        dt = pd.to_datetime(value, errors="coerce")
        if pd.isna(dt):
            return None
        return dt.to_pydatetime()
    except Exception:
        return None


def build_shipment_details(
    shipment_df: pd.DataFrame,
    dashboard_df: pd.DataFrame,
    lead_time: int,
    target_days: int,
    today=None,
) -> dict:
    """Build shipment-wise quantity breakdown with time-aware excess forecasts.
    
    For each shipment:
    1. Calculate days from today until it arrives (available_date)
    2. Deduct daily sales consumption for those days from current stock
    3. Add shipment quantity
    4. Only the amount exceeding required_qty is excess for that shipment
    """
    shipment_df = shipment_df.copy()
    shipment_df["Qty"] = pd.to_numeric(shipment_df["Qty"], errors="coerce").fillna(0).astype(int)
    shipment_df["Loading Date Parsed"] = shipment_df["Loading Date"].apply(parse_date)

    per_shipment = (
        shipment_df.groupby([
            "SKU", "Shipment ID", "Appointment ID", "Loading Date Parsed", "Loading Date"
        ], dropna=False, as_index=False)["Qty"].sum()
    )

    sku_totals = (
        per_shipment.groupby("SKU", as_index=False)["Qty"]
        .agg(total_qty="sum", shipment_count="count", avg_qty="mean")
    )

    today = normalize_day(today)
    shipment_details = []

    for _, sku_row in sku_totals.iterrows():
        sku = sku_row["SKU"]
        sku_metrics = dashboard_df.loc[dashboard_df["SKU"] == sku]
        if not sku_metrics.empty:
            daily_sales = float(sku_metrics["Daily Sales"].iloc[0])
            initial_stock = int(sku_metrics["Current FBA Stock"].iloc[0])
            required_qty = int(sku_metrics["Required Qty"].iloc[0])
        else:
            daily_sales = 0.0
            initial_stock = 0
            required_qty = 0

        shipments = []
        running_stock = initial_stock
        current_day = today
        sku_shipments = per_shipment[per_shipment["SKU"] == sku].sort_values("Loading Date Parsed")

        for _, row in sku_shipments.iterrows():
            loading_date = row["Loading Date Parsed"]
            available_date = (
                loading_date + timedelta(days=lead_time)
                if loading_date is not None
                else None
            )

            if available_date is not None:
                available_day = pd.Timestamp(available_date).normalize()
            else:
                available_day = current_day

            days_until_available = max(0, (available_day - current_day).days)
            consumed = daily_sales * days_until_available
            opening_stock = running_stock
            stock_at_arrival = max(0, running_stock - consumed)

            shipment_qty = int(row["Qty"])
            stock_after_shipment = stock_at_arrival + shipment_qty
            projected_stock = int(round(stock_after_shipment))
            days_cover = 999.0 if daily_sales == 0 else round(projected_stock / daily_sales, 1)

            candidate_excess = max(projected_stock - required_qty, 0)
            shipment_excess_qty = int(candidate_excess if days_cover > target_days else 0)

            shipments.append({
                "id": str(row["Shipment ID"]),
                "qty": shipment_qty,
                "loading_date": (
                    loading_date.strftime("%d %b %Y") if loading_date else str(row["Loading Date"])
                ),
                "appointment_id": str(row["Appointment ID"]),
                "available_date": available_date.strftime("%d %b %Y") if available_date else "Unknown",
                "opening_stock": round(opening_stock, 2),
                "depletion_days": int(days_until_available),
                "expected_sales_before_arrival": round(consumed, 2),
                "stock_before_arrival": int(round(stock_at_arrival)),
                "projected_stock": projected_stock,
                "required_qty": required_qty,
                "days_cover": days_cover,
                "shipment_excess_qty": shipment_excess_qty,
                "status": "Excess" if shipment_excess_qty > 0 else "Within target",
            })

            running_stock = stock_after_shipment
            current_day = available_day

        shipment_details.append({
            "sku": sku,
            "total_qty": int(sku_row["total_qty"]),
            "shipment_count": int(sku_row["shipment_count"]),
            "avg_qty": round(float(sku_row["avg_qty"]), 2),
            "shipments": shipments,
        })

    return {"shipment_details": shipment_details}


# ----------------------------------------------------------
# Routes
# ----------------------------------------------------------
@app.route("/")
def index():
    """Render the dashboard page."""
    return render_template("index.html")


@app.route("/api/analyze", methods=["POST"])
def analyze():
    """
    Main endpoint: receives the 3 uploaded files + settings,
    runs the calculation engine, and returns JSON for the UI.
    """
    try:
        # --- Validate file uploads ---
        for key in ("shipment", "sales", "stock"):
            if key not in request.files:
                return jsonify({"ok": False, "error": f"Missing {key} file."}), 400
            f = request.files[key]
            if f.filename == "":
                return jsonify({"ok": False, "error": f"No {key} file selected."}), 400
            if not allowed_file(f.filename):
                return jsonify({
                    "ok": False,
                    "error": f"Invalid file type for {key}. Use .xlsx or .csv."
                }), 400

        # --- Read settings (with safe defaults) ---
        try:
            transit_days = int(request.form.get("transit_days", 10))
            grn_days     = int(request.form.get("grn_days", 7))
            target_days  = int(request.form.get("target_days", 45))
        except ValueError:
            return jsonify({
                "ok": False,
                "error": "Settings must be whole numbers."
            }), 400

        if min(transit_days, grn_days, target_days) < 0:
            return jsonify({
                "ok": False,
                "error": "Settings cannot be negative."
            }), 400

        # --- Read each file into a DataFrame ---
        shipment_df = normalize_columns(read_any(request.files["shipment"], REQUIRED_COLUMNS["shipment"]))
        sales_df    = normalize_columns(read_any(request.files["sales"], REQUIRED_COLUMNS["sales"]))
        stock_df    = normalize_columns(read_any(request.files["stock"], REQUIRED_COLUMNS["stock"]))

        # --- Validate columns ---
        validate_columns(shipment_df, REQUIRED_COLUMNS["shipment"], "Shipment")
        validate_columns(sales_df,    REQUIRED_COLUMNS["sales"],    "Sales")
        validate_columns(stock_df,    REQUIRED_COLUMNS["stock"],    "Stock")

        # --- Run the engine ---
        result = build_dashboard(
            shipment_df, sales_df, stock_df,
            transit_days, grn_days, target_days,
        )

        kpis   = build_kpis(result)
        charts = build_charts(result)
        shipment_details = build_shipment_details(
            shipment_df,
            result,
            transit_days + grn_days,
            target_days,
        )

        # --- Cache result so download endpoints work ---
        session_id = uuid.uuid4().hex
        LAST_RESULT[session_id] = result

        # Limit cache size — keep only the 10 most recent sessions
        if len(LAST_RESULT) > 10:
            oldest = next(iter(LAST_RESULT))
            LAST_RESULT.pop(oldest, None)

        return jsonify({
            "ok": True,
            "session_id": session_id,
            "kpis": kpis,
            "charts": charts,
            "rows": result.to_dict(orient="records"),
            "shipment_details": shipment_details["shipment_details"],
            "settings": {
                "transit_days": transit_days,
                "grn_days": grn_days,
                "target_days": target_days,
                "lead_time": transit_days + grn_days,
            },
        })

    except ValueError as ve:
        # Friendly business-rule errors (e.g. missing columns)
        return jsonify({"ok": False, "error": str(ve)}), 400
    except Exception as e:
        # Unexpected server error — return a clean message
        return jsonify({
            "ok": False,
            "error": f"Unexpected error: {type(e).__name__}: {e}"
        }), 500


def _dataframe_to_excel(df: pd.DataFrame, sheet_name: str) -> BytesIO:
    """Render any DataFrame to an in-memory .xlsx for download."""
    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    bio.seek(0)
    return bio


@app.route("/api/download/<kind>/<session_id>")
def download(kind: str, session_id: str):
    """
    kind = 'full'   -> entire dashboard
           'excess' -> excess inventory only
           'dead'   -> dead stock only
    """
    df = LAST_RESULT.get(session_id)
    if df is None:
        abort(404, description="Result expired. Please re-run the analysis.")

    if kind == "full":
        out = df.copy()
        name = "FBA_Dashboard_Full.xlsx"
        sheet = "Dashboard"
    elif kind == "excess":
        out = df[df["Excess Qty"] > 0].copy()
        name = "FBA_Excess_Inventory.xlsx"
        sheet = "Excess Inventory"
    elif kind == "dead":
        out = df[df["Status"] == "Dead Stock"].copy()
        name = "FBA_Dead_Stock.xlsx"
        sheet = "Dead Stock"
    else:
        abort(400, description="Unknown report type.")

    bio = _dataframe_to_excel(out, sheet)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = name.replace(".xlsx", f"_{timestamp}.xlsx")

    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ----------------------------------------------------------
# Entry point
# ----------------------------------------------------------
if __name__ == "__main__":
    # Run the development server.
    # Open http://127.0.0.1:5000 in your browser.
    app.run(host="127.0.0.1", port=5000, debug=True)
